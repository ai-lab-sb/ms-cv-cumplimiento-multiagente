"""
Lee archivos Excel (.xlsx, .xls) y CSV, convirtiéndolos a texto plano
para que puedan ser procesados por el LLM como contexto textual.
Gemini no soporta archivos Excel como input multimodal, así que se
convierten a una representación tabular en texto.
"""
import io
import csv


SPREADSHEET_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
}


def es_spreadsheet(mime_type: str) -> bool:
    """Verifica si un mime type corresponde a un archivo de hoja de cálculo."""
    if not mime_type:
        return False
    return mime_type.lower() in SPREADSHEET_MIME_TYPES


def leer_spreadsheet(contenido: bytes, mime_type: str, nombre: str = "") -> str:
    """
    Convierte un archivo Excel o CSV a texto legible.

    Args:
        contenido: Bytes del archivo.
        mime_type: Tipo MIME del archivo.
        nombre: Nombre del archivo (para referencia).

    Returns:
        Texto representando el contenido de la hoja de cálculo.
    """
    mime_lower = mime_type.lower()

    if mime_lower in ("text/csv", "application/csv"):
        return _leer_csv(contenido, nombre)
    else:
        return _leer_excel(contenido, nombre)


def _leer_excel(contenido: bytes, nombre: str) -> str:
    """Lee un archivo Excel (.xlsx/.xls) y lo convierte a texto."""
    try:
        import openpyxl
    except ImportError:
        return f"[ERROR] No se pudo leer el archivo Excel '{nombre}': openpyxl no instalado."

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        resultado = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            resultado.append(f"=== Hoja: {sheet_name} ===")

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                resultado.append("(Hoja vacía)")
                continue

            for row in rows:
                celdas = []
                for cell in row:
                    if cell is None:
                        celdas.append("")
                    else:
                        celdas.append(str(cell).strip())
                if any(c for c in celdas):
                    resultado.append(" | ".join(celdas))

            resultado.append("")

        wb.close()
        texto = "\n".join(resultado)
        if nombre:
            texto = f"Archivo: {nombre}\n\n{texto}"
        return texto

    except Exception as e:
        return f"[ERROR] No se pudo leer el archivo Excel '{nombre}': {e}"


def _leer_csv(contenido: bytes, nombre: str) -> str:
    """Lee un archivo CSV y lo convierte a texto."""
    try:
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                texto_raw = contenido.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            texto_raw = contenido.decode("utf-8", errors="replace")

        reader = csv.reader(io.StringIO(texto_raw))
        resultado = []
        for row in reader:
            celdas = [c.strip() for c in row]
            if any(celdas):
                resultado.append(" | ".join(celdas))

        texto = "\n".join(resultado)
        if nombre:
            texto = f"Archivo: {nombre}\n\n{texto}"
        return texto

    except Exception as e:
        return f"[ERROR] No se pudo leer el archivo CSV '{nombre}': {e}"
